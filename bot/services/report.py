from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from fpdf import FPDF


def generate_excel(path: str, transactions: list[dict], summary: dict, username: str, now: datetime):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ringkasan"
    header_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws.cell(row=1, column=1, value=f"Fintra Monthly Report - {username}").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Periode: {now.strftime('%B %Y')}")
    ws.cell(row=3, column=1, value="")

    headers_r = ["Deskripsi", "Jumlah"]
    for col, h in enumerate(headers_r, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.border = border

    ws.cell(row=5, column=1, value="Total Pemasukan").border = border
    ws.cell(row=5, column=2, value=summary["total_income"]).border = border
    ws.cell(row=6, column=1, value="Total Pengeluaran").border = border
    ws.cell(row=6, column=2, value=summary["total_expense"]).border = border
    ws.cell(row=7, column=1, value="Sisa Saldo").font = Font(bold=True)
    ws.cell(row=7, column=1).border = border
    ws.cell(row=7, column=2, value=summary["total_income"] - summary["total_expense"]).font = Font(bold=True)
    ws.cell(row=7, column=2).border = border

    ws2 = wb.create_sheet("Detail Transaksi")
    headers_d = ["Tanggal", "Tipe", "Nominal", "Kategori", "Catatan"]
    for col, h in enumerate(headers_d, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.border = border

    for i, t in enumerate(transactions, 2):
        ws2.cell(row=i, column=1, value=t["transaction_date"]).border = border
        ws2.cell(row=i, column=2, value=t["type"]).border = border
        ws2.cell(row=i, column=3, value=t["nominal"]).border = border
        ws2.cell(row=i, column=4, value=t["category"]).border = border
        ws2.cell(row=i, column=5, value=t.get("note", "")).border = border

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 15
    ws2.column_dimensions["D"].width = 15
    ws2.column_dimensions["E"].width = 30

    wb.save(path)


class PDFReport(FPDF):
    def header(self):
        self.set_font("Helvetica", "B", 14)
        self.cell(0, 10, "Fintra Monthly Report", align="C", new_x="LMARGIN", new_y="NEXT")
        self.ln(5)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.cell(0, 10, f"Page {self.page_no()}/{{nb}}", align="C")


def generate_pdf(path: str, transactions: list[dict], summary: dict, username: str, now: datetime):
    pdf = PDFReport()
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 8, f"User: {username}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 8, f"Periode: {now.strftime('%B %Y')}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(60, 8, "Deskripsi", border=1)
    pdf.cell(30, 8, "Jumlah", border=1, new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 11)

    pdf.cell(60, 8, "Total Pemasukan", border=1)
    pdf.cell(30, 8, f"Rp{summary['total_income']:,}", border=1, new_x="LMARGIN", new_y="NEXT")
    pdf.cell(60, 8, "Total Pengeluaran", border=1)
    pdf.cell(30, 8, f"Rp{summary['total_expense']:,}", border=1, new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "B", 11)
    balance = summary["total_income"] - summary["total_expense"]
    pdf.cell(60, 8, "Sisa Saldo", border=1)
    pdf.cell(30, 8, f"Rp{balance:,}", border=1, new_x="LMARGIN", new_y="NEXT")

    pdf.ln(10)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Detail Transaksi", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "B", 9)
    col_w = [25, 25, 30, 30, 60]
    headers = ["Tanggal", "Tipe", "Nominal", "Kategori", "Catatan"]
    for w, h in zip(col_w, headers):
        pdf.cell(w, 8, h, border=1)
    pdf.ln()

    pdf.set_font("Helvetica", "", 8)
    for t in transactions:
        pdf.cell(col_w[0], 7, t["transaction_date"], border=1)
        pdf.cell(col_w[1], 7, t["type"], border=1)
        pdf.cell(col_w[2], 7, f"Rp{t['nominal']:,}", border=1)
        pdf.cell(col_w[3], 7, t["category"], border=1)
        note = (t.get("note") or "")[:40]
        pdf.cell(col_w[4], 7, note, border=1)
        pdf.ln()

    pdf.output(path)
